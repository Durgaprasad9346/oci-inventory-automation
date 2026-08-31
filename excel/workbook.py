from datetime import datetime, date, time
from pathlib import Path
import json
import re

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


# ============================================================
# EXCEL SAFE VALUE
# ============================================================

def excel_safe_value(value):
    """
    Convert OCI/Python values into values that openpyxl/Excel
    can safely store.

    Handles:
        - timezone-aware datetime
        - timezone-aware time
        - date
        - dict
        - list
        - tuple
        - set
        - OCI SDK objects
        - primitive values
    """

    if value is None:
        return ""

    # --------------------------------------------------------
    # datetime
    # --------------------------------------------------------

    if isinstance(value, datetime):

        if value.tzinfo is not None:
            return value.replace(
                tzinfo=None
            )

        return value

    # --------------------------------------------------------
    # time
    # --------------------------------------------------------

    if isinstance(value, time):

        if value.tzinfo is not None:
            return value.replace(
                tzinfo=None
            )

        return value

    # --------------------------------------------------------
    # date
    # --------------------------------------------------------

    if isinstance(value, date):
        return value

    # --------------------------------------------------------
    # Dictionary
    # --------------------------------------------------------

    if isinstance(value, dict):

        try:

            return json.dumps(
                value,
                default=serialize_object,
                ensure_ascii=False,
            )

        except Exception:

            return str(value)

    # --------------------------------------------------------
    # List / tuple / set
    # --------------------------------------------------------

    if isinstance(
        value,
        (list, tuple, set),
    ):

        try:

            return json.dumps(
                list(value),
                default=serialize_object,
                ensure_ascii=False,
            )

        except Exception:

            return str(value)

    # --------------------------------------------------------
    # OCI SDK objects
    # --------------------------------------------------------

    if not isinstance(
        value,
        (
            str,
            int,
            float,
            bool,
        ),
    ):

        # Some OCI SDK objects contain a to_dict()
        # method.

        if hasattr(
            value,
            "to_dict",
        ):

            try:

                converted = value.to_dict()

                return json.dumps(
                    converted,
                    default=serialize_object,
                    ensure_ascii=False,
                )

            except Exception:
                pass

        # Some SDK objects expose __dict__.

        if hasattr(
            value,
            "__dict__",
        ):

            try:

                converted = {}

                for key, item in (
                    value.__dict__.items()
                ):

                    if key.startswith("_"):
                        continue

                    converted[key] = item

                return json.dumps(
                    converted,
                    default=serialize_object,
                    ensure_ascii=False,
                )

            except Exception:
                pass

    # --------------------------------------------------------
    # Primitive values
    # --------------------------------------------------------

    return value


# ============================================================
# SERIALIZE COMPLEX OBJECTS
# ============================================================

def serialize_object(value):
    """
    JSON serializer for OCI SDK objects and nested values.
    """

    if value is None:
        return None

    if isinstance(value, datetime):

        if value.tzinfo is not None:
            value = value.replace(
                tzinfo=None
            )

        return value.isoformat(
            sep=" "
        )

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, time):

        if value.tzinfo is not None:
            value = value.replace(
                tzinfo=None
            )

        return value.isoformat()

    if isinstance(
        value,
        (list, tuple, set),
    ):

        return list(value)

    if isinstance(value, dict):
        return value

    if hasattr(
        value,
        "to_dict",
    ):

        try:
            return value.to_dict()
        except Exception:
            pass

    if hasattr(
        value,
        "__dict__",
    ):

        try:

            return {
                key: item
                for key, item in value.__dict__.items()
                if not key.startswith("_")
            }

        except Exception:
            pass

    return str(value)


# ============================================================
# SAFE STRING
# ============================================================

def safe_string(value):

    if value is None:
        return ""

    if isinstance(
        value,
        (
            dict,
            list,
            tuple,
            set,
        ),
    ):

        try:

            return json.dumps(
                value,
                default=serialize_object,
                ensure_ascii=False,
            )

        except Exception:

            return str(value)

    return str(value)


# ============================================================
# FLATTEN DICTIONARY
# ============================================================

def flatten_dict(
    data,
    prefix="",
):
    """
    Flatten nested dictionaries.

    Example:

        {
            "maxlife": {
                "env": "prod",
                "project": "abc"
            }
        }

    becomes:

        maxlife.env = prod
        maxlife.project = abc
    """

    result = {}

    if not isinstance(
        data,
        dict,
    ):

        return result

    for key, value in data.items():

        key = str(key)

        if prefix:

            new_key = (
                f"{prefix}.{key}"
            )

        else:

            new_key = key

        if isinstance(
            value,
            dict,
        ):

            nested = flatten_dict(
                value,
                new_key,
            )

            result.update(
                nested
            )

        else:

            result[
                new_key
            ] = excel_safe_value(
                value
            )

    return result


# ============================================================
# GET FIELD
# ============================================================

def get_field(
    obj,
    *names,
    default=None,
):
    """
    Get a field from either a dictionary or object.
    """

    for name in names:

        if isinstance(
            obj,
            dict,
        ):

            if name in obj:
                return obj[name]

        else:

            if hasattr(
                obj,
                name,
            ):

                try:

                    value = getattr(
                        obj,
                        name,
                    )

                    if value is not None:
                        return value

                except Exception:
                    pass

    return default


# ============================================================
# RESOURCE TO DICT
# ============================================================

def resource_to_dict(resource):
    """
    Convert Resource objects into dictionaries.
    """

    if isinstance(
        resource,
        dict,
    ):

        return dict(
            resource
        )

    data = {}

    fields = [

        "service",
        "service_name",

        "resource_type",
        "resourceType",

        "name",
        "display_name",
        "displayName",

        "ocid",
        "id",
        "identifier",

        "compartment_id",
        "compartmentId",

        "compartment_name",
        "compartmentName",

        "region",

        "availability_domain",
        "availabilityDomain",

        "state",

        "lifecycle_state",
        "lifecycleState",

        "lifecycle_details",
        "lifecycleDetails",

        "time_created",
        "timeCreated",

        "creation_date",
        "created_at",

        "defined_tags",
        "definedTags",

        "freeform_tags",
        "freeformTags",

        "details",

        "additional_details",
        "additionalDetails",
    ]

    # --------------------------------------------------------
    # Known fields
    # --------------------------------------------------------

    for field in fields:

        if hasattr(
            resource,
            field,
        ):

            try:

                value = getattr(
                    resource,
                    field,
                )

                if value is not None:

                    data[
                        field
                    ] = value

            except Exception:
                pass

    # --------------------------------------------------------
    # Object __dict__
    # --------------------------------------------------------

    if hasattr(
        resource,
        "__dict__",
    ):

        try:

            for key, value in (
                resource.__dict__.items()
            ):

                if key.startswith("_"):
                    continue

                if key not in data:

                    data[key] = value

        except Exception:
            pass

    return data


# ============================================================
# DEFINED TAGS
# ============================================================

def extract_defined_tags(
    data,
):

    tags = (
        data.get(
            "defined_tags"
        )
        or data.get(
            "definedTags"
        )
        or {}
    )

    if not isinstance(
        tags,
        dict,
    ):

        return {}

    return flatten_dict(
        tags
    )


# ============================================================
# FREEFORM TAGS
# ============================================================

def extract_freeform_tags(
    data,
):

    tags = (
        data.get(
            "freeform_tags"
        )
        or data.get(
            "freeformTags"
        )
        or {}
    )

    if not isinstance(
        tags,
        dict,
    ):

        return {}

    return flatten_dict(
        tags
    )


# ============================================================
# TAG FILTER
# ============================================================

def include_tag(
    tag_name,
):

    if not tag_name:
        return False

    tag_name = str(
        tag_name
    )

    # Keep Schedule tags in the main resource details,
    # but do not create individual Schedule:* columns.

    if tag_name.lower().startswith(
        "schedule:"
    ):

        return False

    return True


# ============================================================
# CREATION DATE
# ============================================================

def get_creation_date(
    data,
):

    value = (
        data.get(
            "time_created"
        )
        or data.get(
            "timeCreated"
        )
        or data.get(
            "creation_date"
        )
        or data.get(
            "created_at"
        )
    )

    return excel_safe_value(
        value
    )


# ============================================================
# STANDARD COLUMNS
# ============================================================

STANDARD_COLUMNS = [

    "Service",

    "Resource Type",

    "Resource Name",

    "Resource OCID",

    "Compartment Name",

    "Compartment OCID",

    "Region",

    "Availability Domain",

    "Lifecycle State",

    "Lifecycle Details",

    "Creation Date",
]


# ============================================================
# BUILD RESOURCE ROW
# ============================================================

def build_resource_row(
    resource,
):

    data = resource_to_dict(
        resource
    )

    row = {}

    # --------------------------------------------------------
    # Service
    # --------------------------------------------------------

    row["Service"] = (
        data.get(
            "service"
        )
        or data.get(
            "service_name"
        )
        or ""
    )

    # --------------------------------------------------------
    # Resource Type
    # --------------------------------------------------------

    row["Resource Type"] = (
        data.get(
            "resource_type"
        )
        or data.get(
            "resourceType"
        )
        or ""
    )

    # --------------------------------------------------------
    # Name
    # --------------------------------------------------------

    row["Resource Name"] = (
        data.get(
            "name"
        )
        or data.get(
            "display_name"
        )
        or data.get(
            "displayName"
        )
        or ""
    )

    # --------------------------------------------------------
    # OCID
    # --------------------------------------------------------

    row["Resource OCID"] = (
        data.get(
            "ocid"
        )
        or data.get(
            "identifier"
        )
        or data.get(
            "id"
        )
        or ""
    )

    # --------------------------------------------------------
    # Compartment
    # --------------------------------------------------------

    row["Compartment Name"] = (
        data.get(
            "compartment_name"
        )
        or data.get(
            "compartmentName"
        )
        or ""
    )

    row["Compartment OCID"] = (
        data.get(
            "compartment_id"
        )
        or data.get(
            "compartmentId"
        )
        or ""
    )

    # --------------------------------------------------------
    # Region
    # --------------------------------------------------------

    row["Region"] = (
        data.get(
            "region"
        )
        or ""
    )

    # --------------------------------------------------------
    # Availability Domain
    # --------------------------------------------------------

    row["Availability Domain"] = (
        data.get(
            "availability_domain"
        )
        or data.get(
            "availabilityDomain"
        )
        or ""
    )

    # --------------------------------------------------------
    # Lifecycle State
    # --------------------------------------------------------

    row["Lifecycle State"] = (
        data.get(
            "state"
        )
        or data.get(
            "lifecycle_state"
        )
        or data.get(
            "lifecycleState"
        )
        or ""
    )

    # --------------------------------------------------------
    # Lifecycle Details
    # --------------------------------------------------------

    row["Lifecycle Details"] = (
        data.get(
            "lifecycle_details"
        )
        or data.get(
            "lifecycleDetails"
        )
        or ""
    )

    # --------------------------------------------------------
    # Creation Date
    # --------------------------------------------------------

    row["Creation Date"] = (
        get_creation_date(
            data
        )
    )

    # ========================================================
    # DEFINED TAGS
    # ========================================================

    defined_tags = (
        extract_defined_tags(
            data
        )
    )

    for tag_name, tag_value in (
        defined_tags.items()
    ):

        if include_tag(
            tag_name
        ):

            row[
                f"Tag: {tag_name}"
            ] = excel_safe_value(
                tag_value
            )

    # ========================================================
    # FREEFORM TAGS
    # ========================================================

    freeform_tags = (
        extract_freeform_tags(
            data
        )
    )

    for tag_name, tag_value in (
        freeform_tags.items()
    ):

        if include_tag(
            tag_name
        ):

            row[
                f"Freeform Tag: {tag_name}"
            ] = excel_safe_value(
                tag_value
            )

    # ========================================================
    # RESOURCE DETAILS
    # ========================================================

    details = (
        data.get(
            "details"
        )
        or data.get(
            "additional_details"
        )
        or data.get(
            "additionalDetails"
        )
        or {}
    )

    if isinstance(
        details,
        dict,
    ):

        flattened_details = (
            flatten_dict(
                details,
                "Details",
            )
        )

        for key, value in (
            flattened_details.items()
        ):

            if key not in row:

                row[
                    key
                ] = excel_safe_value(
                    value
                )

    return row


# ============================================================
# GET COLUMNS
# ============================================================

def get_columns(
    rows,
):

    columns = list(
        STANDARD_COLUMNS
    )

    tag_columns = set()

    other_columns = set()

    for row in rows:

        for key in row.keys():

            if key in columns:
                continue

            if (
                str(key).startswith(
                    "Tag:"
                )
                or str(key).startswith(
                    "Freeform Tag:"
                )
            ):

                tag_columns.add(
                    key
                )

            else:

                other_columns.add(
                    key
                )

    columns.extend(
        sorted(
            tag_columns
        )
    )

    columns.extend(
        sorted(
            other_columns
        )
    )

    return columns


# ============================================================
# SANITIZE SHEET NAME
# ============================================================

def sanitize_sheet_name(
    name,
):

    if not name:
        name = "Resources"

    name = str(
        name
    )

    name = re.sub(
        r"[\[\]\:\*\?\/\\]",
        "_",
        name,
    )

    name = name.strip()

    if not name:
        name = "Resources"

    return name[:31]


# ============================================================
# UNIQUE SHEET NAME
# ============================================================

def get_unique_sheet_name(
    workbook,
    name,
):

    base = sanitize_sheet_name(
        name
    )

    if base not in workbook.sheetnames:
        return base

    counter = 2

    while True:

        suffix = (
            f"_{counter}"
        )

        candidate = (
            base[
                :31 - len(suffix)
            ]
            + suffix
        )

        if candidate not in (
            workbook.sheetnames
        ):

            return candidate

        counter += 1


# ============================================================
# WRITE RESOURCE SHEET
# ============================================================

def write_resource_sheet(
    workbook,
    sheet_name,
    resources,
):

    if not resources:
        return None

    sheet_name = (
        get_unique_sheet_name(
            workbook,
            sheet_name,
        )
    )

    ws = workbook.create_sheet(
        title=sheet_name
    )

    rows = []

    # --------------------------------------------------------
    # Convert resources
    # --------------------------------------------------------

    for resource in resources:

        try:

            rows.append(
                build_resource_row(
                    resource
                )
            )

        except Exception as error:

            print(
                "WARNING: Could not "
                "process resource for "
                f"Excel: {error}"
            )

    if not rows:
        return ws

    columns = get_columns(
        rows
    )

    # ========================================================
    # HEADER
    # ========================================================

    for col_num, column in enumerate(
        columns,
        start=1,
    ):

        cell = ws.cell(
            row=1,
            column=col_num,
            value=column,
        )

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    # ========================================================
    # DATA
    # ========================================================

    for row_num, row_data in enumerate(
        rows,
        start=2,
    ):

        for col_num, column in enumerate(
            columns,
            start=1,
        ):

            value = row_data.get(
                column,
                "",
            )

            # ------------------------------------------------
            # FINAL VALUE CONVERSION
            # ------------------------------------------------

            value = excel_safe_value(
                value
            )

            cell = ws.cell(
                row=row_num,
                column=col_num,
            )

            cell.value = value

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

            # ------------------------------------------------
            # Creation Date format
            # ------------------------------------------------

            if (
                column == "Creation Date"
                and isinstance(
                    value,
                    datetime,
                )
            ):

                cell.number_format = (
                    "yyyy-mm-dd hh:mm:ss"
                )

    # ========================================================
    # FREEZE
    # ========================================================

    ws.freeze_panes = "A2"

    # ========================================================
    # FILTER
    # ========================================================

    ws.auto_filter.ref = (
        ws.dimensions
    )

    # ========================================================
    # TABLE
    # ========================================================

    if ws.max_row >= 2:

        table_base = re.sub(
            r"[^A-Za-z0-9_]",
            "",
            sheet_name,
        )

        if not table_base:

            table_base = (
                "Resources"
            )

        table_name = (
            "tbl_"
            + table_base[:20]
        )

        existing_names = set()

        for existing_ws in (
            workbook.worksheets
        ):

            for existing_table in (
                existing_ws.tables.keys()
            ):

                existing_names.add(
                    existing_table
                )

        original_name = (
            table_name
        )

        counter = 2

        while table_name in (
            existing_names
        ):

            table_name = (
                original_name
                + str(counter)
            )

            counter += 1

        table_ref = (
            f"A1:"
            f"{get_column_letter(ws.max_column)}"
            f"{ws.max_row}"
        )

        table = Table(
            displayName=table_name,
            ref=table_ref,
        )

        table_style = (
            TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
        )

        table.tableStyleInfo = (
            table_style
        )

        ws.add_table(
            table
        )

    # ========================================================
    # COLUMN WIDTH
    # ========================================================

    for col_num in range(
        1,
        ws.max_column + 1,
    ):

        letter = (
            get_column_letter(
                col_num
            )
        )

        max_length = 0

        for cell in ws[
            letter
        ]:

            if cell.value is None:
                continue

            length = len(
                str(
                    cell.value
                )
            )

            if length > max_length:
                max_length = length

        width = min(
            max(
                max_length + 2,
                12,
            ),
            60,
        )

        header = ws.cell(
            row=1,
            column=col_num,
        ).value

        if header in (
            "Resource OCID",
            "Compartment OCID",
        ):

            width = 55

        ws.column_dimensions[
            letter
        ].width = width

    ws.row_dimensions[
        1
    ].height = 30

    return ws


# ============================================================
# FLATTEN resources_by_service
# ============================================================

def flatten_resources_by_service(
    resources_by_service,
):

    resources = []

    if resources_by_service is None:
        return resources

    # --------------------------------------------------------
    # Already a list
    # --------------------------------------------------------

    if isinstance(
        resources_by_service,
        list,
    ):

        return list(
            resources_by_service
        )

    # --------------------------------------------------------
    # Single object
    # --------------------------------------------------------

    if not isinstance(
        resources_by_service,
        dict,
    ):

        return [
            resources_by_service
        ]

    # --------------------------------------------------------
    # Dictionary
    # --------------------------------------------------------

    for service_value in (
        resources_by_service.values()
    ):

        if service_value is None:
            continue

        # Service -> list
        if isinstance(
            service_value,
            list,
        ):

            resources.extend(
                service_value
            )

            continue

        # Service -> dictionary
        if isinstance(
            service_value,
            dict,
        ):

            for resource_value in (
                service_value.values()
            ):

                if resource_value is None:
                    continue

                if isinstance(
                    resource_value,
                    list,
                ):

                    resources.extend(
                        resource_value
                    )

                else:

                    resources.append(
                        resource_value
                    )

            continue

        # Single resource
        resources.append(
            service_value
        )

    return resources


# ============================================================
# GROUP RESOURCES
# ============================================================

def group_resources(
    resources,
):

    grouped = {}

    for resource in resources:

        data = resource_to_dict(
            resource
        )

        service = (
            data.get(
                "service"
            )
            or data.get(
                "service_name"
            )
            or "Other"
        )

        resource_type = (
            data.get(
                "resource_type"
            )
            or data.get(
                "resourceType"
            )
            or "Resource"
        )

        key = (
            str(service),
            str(resource_type),
        )

        grouped.setdefault(
            key,
            [],
        ).append(
            resource
        )

    return grouped


# ============================================================
# SUMMARY SHEET
# ============================================================

def create_summary_sheet(
    workbook,
    resources,
):

    ws = workbook.create_sheet(
        title="Summary",
        index=0,
    )

    # --------------------------------------------------------
    # Title
    # --------------------------------------------------------

    ws["A1"] = (
        "OCI Tenancy Resource Inventory"
    )

    ws["A1"].font = Font(
        bold=True,
        size=16,
    )

    ws.merge_cells(
        "A1:E1"
    )

    # --------------------------------------------------------
    # Report generated
    # --------------------------------------------------------

    ws["A3"] = (
        "Report Generated"
    )

    ws["B3"] = excel_safe_value(
        datetime.now()
    )

    ws["B3"].number_format = (
        "yyyy-mm-dd hh:mm:ss"
    )

    # --------------------------------------------------------
    # Total
    # --------------------------------------------------------

    ws["A5"] = (
        "Total Resources"
    )

    ws["B5"] = len(
        resources
    )

    # --------------------------------------------------------
    # Counts
    # --------------------------------------------------------

    service_counts = {}

    resource_type_counts = {}

    for resource in resources:

        data = resource_to_dict(
            resource
        )

        service = (
            data.get(
                "service"
            )
            or data.get(
                "service_name"
            )
            or "Unknown"
        )

        resource_type = (
            data.get(
                "resource_type"
            )
            or data.get(
                "resourceType"
            )
            or "Unknown"
        )

        service = str(
            service
        )

        resource_type = str(
            resource_type
        )

        service_counts[
            service
        ] = (
            service_counts.get(
                service,
                0,
            )
            + 1
        )

        resource_type_counts[
            resource_type
        ] = (
            resource_type_counts.get(
                resource_type,
                0,
            )
            + 1
        )

    # ========================================================
    # RESOURCE TYPE SUMMARY
    # ========================================================

    ws["A7"] = (
        "Resource Type"
    )

    ws["B7"] = (
        "Count"
    )

    ws["A7"].font = Font(
        bold=True
    )

    ws["B7"].font = Font(
        bold=True
    )

    row = 8

    for resource_type in sorted(
        resource_type_counts
    ):

        ws.cell(
            row=row,
            column=1,
            value=resource_type,
        )

        ws.cell(
            row=row,
            column=2,
            value=resource_type_counts[
                resource_type
            ],
        )

        row += 1

    # ========================================================
    # SERVICE SUMMARY
    # ========================================================

    ws["D7"] = (
        "Service"
    )

    ws["E7"] = (
        "Count"
    )

    ws["D7"].font = Font(
        bold=True
    )

    ws["E7"].font = Font(
        bold=True
    )

    row = 8

    for service in sorted(
        service_counts
    ):

        ws.cell(
            row=row,
            column=4,
            value=service,
        )

        ws.cell(
            row=row,
            column=5,
            value=service_counts[
                service
            ],
        )

        row += 1

    # --------------------------------------------------------
    # Formatting
    # --------------------------------------------------------

    ws.freeze_panes = "A8"

    ws.column_dimensions[
        "A"
    ].width = 35

    ws.column_dimensions[
        "B"
    ].width = 15

    ws.column_dimensions[
        "C"
    ].width = 5

    ws.column_dimensions[
        "D"
    ].width = 35

    ws.column_dimensions[
        "E"
    ].width = 15

    return ws


# ============================================================
# CREATE INVENTORY WORKBOOK
# ============================================================

def create_inventory_workbook(
    resources=None,
    resources_by_service=None,
    output_file=None,
    **kwargs,
):
    """
    Create the complete OCI inventory workbook.

    Compatible with the existing main.py interface:

        create_inventory_workbook(
            resources_by_service=...
        )

    Also supports:

        create_inventory_workbook(
            resources=...
        )
    """

    # ========================================================
    # GET RESOURCE LIST
    # ========================================================

    if resources_by_service is not None:

        resources = (
            flatten_resources_by_service(
                resources_by_service
            )
        )

    elif resources is None:

        resources = []

    else:

        resources = list(
            resources
        )

    # ========================================================
    # OUTPUT FILE
    # ========================================================

    if output_file is None:

        output_file = (
            kwargs.get(
                "output_path"
            )
            or kwargs.get(
                "filename"
            )
            or kwargs.get(
                "file_path"
            )
            or kwargs.get(
                "report_file"
            )
        )

    # ========================================================
    # CREATE WORKBOOK
    # ========================================================

    workbook = Workbook()

    default_sheet = (
        workbook.active
    )

    workbook.remove(
        default_sheet
    )

    # ========================================================
    # SUMMARY
    # ========================================================

    create_summary_sheet(
        workbook,
        resources,
    )

    # ========================================================
    # GROUP
    # ========================================================

    grouped = group_resources(
        resources
    )

    # ========================================================
    # CREATE RESOURCE SHEETS
    # ========================================================

    for (
        service,
        resource_type,
    ), resource_list in sorted(
        grouped.items(),
        key=lambda item: (
            str(item[0][0]),
            str(item[0][1]),
        ),
    ):

        write_resource_sheet(
            workbook,
            resource_type,
            resource_list,
        )

    # ========================================================
    # FINAL EXCEL SAFETY PASS
    # ========================================================

    # IMPORTANT:
    #
    # Every value is passed through excel_safe_value().
    #
    # This prevents:
    #
    #   timezone-aware datetime errors
    #
    # and:
    #
    #   Cannot convert {...} to Excel
    #
    # errors.
    # ========================================================

    for ws in workbook.worksheets:

        for row in ws.iter_rows():

            for cell in row:

                if cell.value is not None:

                    cell.value = (
                        excel_safe_value(
                            cell.value
                        )
                    )

    # ========================================================
    # SAVE
    # ========================================================

    if output_file:

        output_path = Path(
            output_file
        )

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook.save(
            output_path
        )

        print(
            "Inventory workbook created: "
            f"{output_path}"
        )

    return workbook


# ============================================================
# BACKWARD COMPATIBILITY
# ============================================================

def create_workbook(
    resources=None,
    resources_by_service=None,
    output_file=None,
    **kwargs,
):

    return create_inventory_workbook(
        resources=resources,
        resources_by_service=resources_by_service,
        output_file=output_file,
        **kwargs,
    )
